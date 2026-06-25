"""Genera file demo e template pronti all'uso."""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

def header_style(ws, headers, row=1, bg="FF1F3A5F"):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

def set_widths(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

# ── 1. ESTRATTO CONTO DEMO ──────────────────────────────────────────────────
def make_estratto():
    wb = Workbook(); ws = wb.active; ws.title = "Movimenti"
    headers = ["Data Operazione", "Descrizione", "Dare", "Avere", "Saldo"]
    header_style(ws, headers)
    set_widths(ws, [14, 50, 12, 12, 14])
    rows = [
        ["02/06/2025", "Bonifico entrata – ALFA SRL rif. FATT-2025-041",          None,   12200.00, 112200.00],
        ["05/06/2025", "Pag. fornitore BETA FORNITURE – fatt. FRN-0088",          5800.00, None,    106400.00],
        ["10/06/2025", "Accredito GAMMA TECH fattura n. 2025-042",                None,    3680.00, 110080.00],
        ["12/06/2025", "Addebito fatt. FRN-0089 DELTA LOGISTICA",                 2100.00, None,   107980.00],
        ["15/06/2025", "Canone servizi bancari giugno",                            45.00,   None,   107935.00],
        ["18/06/2025", "Accredito EPSILON SPA – inv 2025-043",                    None,    9500.00, 117435.00],
        ["20/06/2025", "Pagamento FRN-0090 – ZETA SERVIZI SRL",                   1250.00, None,   116185.00],
        ["22/06/2025", "Bonifico entrata ETA GROUP fatt 2025-044",                None,    7300.00, 123485.00],
        ["23/06/2025", "Rimborso spese dipendente Rossi",                          320.00,  None,  123165.00],
        ["25/06/2025", "Pagamento non identificato – causale mancante",            980.00,  None,  122185.00],
        ["27/06/2025", "Accredito THETA CONSULTING – fattura 2025-045",           None,    4400.00, 126585.00],
        ["28/06/2025", "Pag. parziale fornitore IOTA IMPIANTI FRN-0091",          2800.00, None,  123785.00],
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            if j == 1: c.number_format = "DD/MM/YYYY"
            if j in (3,4,5): c.number_format = '#,##0.00 €'
    wb.save(BASE / "input/estratti/estratto_giugno2025_DEMO.xlsx")
    print("  ✓ estratto_giugno2025_DEMO.xlsx")

# ── 2. PARTITARIO FORNITORI DEMO ────────────────────────────────────────────
def make_fornitori():
    wb = Workbook(); ws = wb.active; ws.title = "Partitario Fornitori"
    headers = ["Numero Documento", "Tipo", "Controparte", "Data Documento",
               "Data Scadenza", "Importo", "Stato", "Contestata", "Motivo Contestazione", "Note"]
    header_style(ws, headers)
    set_widths(ws, [16,10,24,14,14,12,14,10,30,24])
    rows = [
        ["FRN-0088","Ricevuta","Beta Forniture SRL","01/05/2025","05/06/2025",5800.00,"Aperta","No","",""],
        ["FRN-0089","Ricevuta","Delta Logistica SpA","03/05/2025","12/06/2025",2100.00,"Aperta","No","",""],
        ["FRN-0090","Ricevuta","Zeta Servizi SRL","10/05/2025","20/06/2025",1250.00,"Aperta","No","",""],
        ["FRN-0091","Ricevuta","Iota Impianti Srl","15/05/2025","28/06/2025",3300.00,"Aperta","Sì","Importo non concordato – attesa NC","Verificare con resp. acquisti"],
        ["FRN-0092","Ricevuta","Kappa Consulting","20/05/2025","30/06/2025",2750.00,"Aperta","No","","Da verificare precheck"],
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            if j in (4,5): c.number_format = "DD/MM/YYYY"
            if j == 6: c.number_format = '#,##0.00 €'
    wb.save(BASE / "input/partitari/partitario_fornitori_DEMO.xlsx")
    print("  ✓ partitario_fornitori_DEMO.xlsx")

# ── 3. PARTITARIO CLIENTI DEMO ──────────────────────────────────────────────
def make_clienti():
    wb = Workbook(); ws = wb.active; ws.title = "Partitario Clienti"
    headers = ["Numero Documento", "Tipo", "Controparte", "Data Documento",
               "Data Scadenza", "Importo", "Stato", "Contestata", "Motivo Contestazione", "Note"]
    header_style(ws, headers)
    set_widths(ws, [16,10,24,14,14,12,14,10,30,24])
    rows = [
        ["FATT-2025-041","Emessa","Alfa SRL","01/05/2025","02/06/2025",12200.00,"Aperta","No","",""],
        ["FATT-2025-042","Emessa","Gamma Tech SpA","05/05/2025","10/06/2025",3680.00,"Aperta","No","",""],
        ["FATT-2025-043","Emessa","Epsilon SpA","08/05/2025","18/06/2025",9500.00,"Aperta","No","",""],
        ["FATT-2025-044","Emessa","Eta Group Srl","12/05/2025","22/06/2025",7300.00,"Aperta","No","",""],
        ["FATT-2025-045","Emessa","Theta Consulting","18/05/2025","27/06/2025",4400.00,"Aperta","No","",""],
        ["FATT-2025-046","Emessa","Lambda Corp","25/05/2025","30/07/2025",6800.00,"Aperta","Sì","Cliente contesta servizio non erogato","In attesa risposta cliente"],
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            if j in (4,5): c.number_format = "DD/MM/YYYY"
            if j == 6: c.number_format = '#,##0.00 €'
    wb.save(BASE / "input/partitari/partitario_clienti_DEMO.xlsx")
    print("  ✓ partitario_clienti_DEMO.xlsx")

# ── 4. CONTESTAZIONI DEMO ───────────────────────────────────────────────────
def make_contestazioni():
    wb = Workbook(); ws = wb.active; ws.title = "Contestazioni"
    headers = ["Numero Documento","Stato Contestazione","Data Contestazione","Motivo","Responsabile","Note Interne"]
    header_style(ws, headers)
    set_widths(ws, [16,22,16,36,18,30])
    rows = [
        ["FRN-0091","contestata_da_noi","10/06/2025","Importo superiore all'ordine (#PO-447)","Mario Bianchi","Attesa nota credito da fornitore"],
        ["FATT-2025-046","contestata_dal_cliente","15/06/2025","Cliente nega ricezione servizio","Giulia Rossi","Inviata documentazione il 18/06"],
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            if j == 3: c.number_format = "DD/MM/YYYY"
    wb.save(BASE / "input/contestazioni_DEMO.xlsx")
    print("  ✓ contestazioni_DEMO.xlsx")

# ── 5. TEMPLATE VUOTI ────────────────────────────────────────────────────────
def make_templates():
    templates = {
        "TEMPLATE_estratto_conto.xlsx": {
            "sheet": "Movimenti",
            "headers": ["Data Operazione","Descrizione","Dare","Avere","Saldo"],
            "widths": [14,50,12,12,14],
            "note": "Colonne obbligatorie: Data Operazione, Descrizione, Dare O Avere. Saldo opzionale.",
        },
        "TEMPLATE_partitario.xlsx": {
            "sheet": "Partitario",
            "headers": ["Numero Documento","Tipo","Controparte","Data Documento",
                        "Data Scadenza","Importo","Stato","Contestata","Motivo Contestazione","Note"],
            "widths": [16,10,24,14,14,12,14,10,30,24],
            "note": "Tipo: Emessa (fatture clienti) o Ricevuta (fatture fornitori). Contestata: Sì / No.",
        },
        "TEMPLATE_contestazioni.xlsx": {
            "sheet": "Contestazioni",
            "headers": ["Numero Documento","Stato Contestazione","Data Contestazione","Motivo","Responsabile","Note Interne"],
            "widths": [16,22,16,36,18,30],
            "note": "Stato: contestata_da_noi | contestata_dal_cliente | in_attesa_nota_credito | risolta | aperta",
        },
    }
    for fname, cfg in templates.items():
        wb = Workbook(); ws = wb.active; ws.title = cfg["sheet"]
        header_style(ws, cfg["headers"])
        set_widths(ws, cfg["widths"])
        # riga nota
        note_row = len(cfg["headers"]) + 2
        ws.cell(row=3, column=1, value=f"ℹ️  {cfg['note']}").font = Font(italic=True, color="FF888888", size=9)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(cfg["headers"]))
        wb.save(BASE / f"input/{fname}")
        print(f"  ✓ {fname}")

if __name__ == "__main__":
    print("\nGenerazione file demo e template...\n")
    make_estratto()
    make_fornitori()
    make_clienti()
    make_contestazioni()
    make_templates()
    print("\nFatto! Tutti i file sono in input/\n")
